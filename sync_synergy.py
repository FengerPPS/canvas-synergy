




#import pytest
from lib2to3.pgen2 import driver
import time
import json
import csv
import re
import glob
import datetime
import pathlib
import os
import socket
from pathlib import Path

from argparse import ArgumentParser
from selenium import webdriver
from selenium.webdriver.common.by import By

from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.alert import Alert

#from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
from selenium.common.exceptions import UnexpectedAlertPresentException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import ElementNotInteractableException
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support import expected_conditions as EC

from contextlib import contextmanager
from selenium.webdriver.support.expected_conditions import staleness_of



from openpyxl import load_workbook
from openpyxl import Workbook
import openpyxl
import numpy as np


from fuzzywuzzy import fuzz

import warnings
import configparser
import pwinput
import getpass
from cryptography.fernet import Fernet



SCORE_SAVE_DELAY = 7 #Seconds

def deletefile(myfile):

    ## Try to delete the file ##
    try:
        file_to_rem = pathlib.Path(myfile)
        file_to_rem.unlink()
    
    except FileNotFoundError as e:  ## if failed, report it back to the user ##
        print ("Error: %s - %s." % (myfile, e.strerror))

def handleAlert(driver):
    try:
        alert = driver.switch_to.alert
        alert.accept()
    except (NoAlertPresentException) as err:
        print("No Alert")

def click_ni_element(el, max_attempts=10):
    attempt = 1
    while True:
        try:
            #return self.browser.is_text_present(text)
            el.click()
            return
        except (ElementNotInteractableException, ElementClickInterceptedException) as err:
        
            if attempt == max_attempts:
                raise
            print("Not Clickable or Interactable... " + str(attempt))
            print(err)
            print(err.args)
            attempt += 1
            time.sleep(.5)

def check_exists_by_xpath(driver, xpath):
    try:
        driver.find_element_by_xpath(xpath)
    except NoSuchElementException:
        return False
    return True


def click_stale_element(el, max_attempts=10):
    attempt = 1
    while True:
        try:
            #return self.browser.is_text_present(text)
            el.click()
            return
        except StaleElementReferenceException:
            if max_attempts == 0:
                return
            elif attempt == max_attempts:
                raise
            print("Stale... " + str(attempt))
            attempt += 1
            time.sleep(1)

def click_stale_element_xpath(driver, xpath, max_attempts=10):
    attempt = 1
    while True:
        try:
            #return self.browser.is_text_present(text)
            driver.find_element(By.XPATH,xpath).click()
            return
        except StaleElementReferenceException:
            if attempt == max_attempts:
                raise
            print("Stale... " + str(attempt))
            attempt += 1
            time.sleep(1)

def click_stale_element_id(driver, idtext, max_attempts=10):
    attempt = 1
    while True:
        try:
            #return self.browser.is_text_present(text)
            driver.find_element(By.ID,idtext).click()
            return
        except StaleElementReferenceException:
            if attempt == max_attempts:
                raise
            print("Stale... " + str(attempt))
            attempt += 1
            time.sleep(1)


def wait_for_element_to_load(driver, element_id, max_delay=10):
    try:
        myElem = WebDriverWait(driver, max_delay).until(EC.presence_of_element_located((By.ID, element_id)))
        #print( "Element is ready!")
    except TimeoutException:
        print( "Element loading took too much time!")

# def page_has_loaded_id(driver):

    # try:
        # new_page = driver.find_element(By.TAG_NAME,'html')
        # return new_page.id != old_page.id
    # except NoSuchElementException:
        # return False

#http://www.obeythetestinggoat.com/how-to-get-selenium-to-wait-for-page-load-after-a-click.html
@contextmanager
def wait_for_page_load(driver, timeout=30):
    old_page = driver.find_element(By.TAG_NAME,'html')
    yield
    WebDriverWait(driver, timeout).until(
        staleness_of(old_page)
    )



def xstr(s):
    if s is None:
        return ''
    return str(s)

def addscores(driver, csvfilename, section, assignmentstoskip ):
    #driver.execute_script("document.body.style.zoom='75%'")
    num_updates = 0
    print("\nAdding scores for " + section + "\n")
    #First two rows of Canvas csv export are headings
    first_csv_datarow = 3 
    canvas_headings = []

    with open(csvfilename, newline='') as csvfile:
        data = list(csv.reader(csvfile, delimiter=","))
    #Transpose list
    #datatrans = list(map(list, zip(*data)))

    num_csv_rows = len(data) + 1
    num_csv_cols = len(data[0])
    
    time.sleep(.5)
    student_elements = driver.find_elements(By.XPATH,"//*[contains(@id,'_lbl_studentLastFirst')]")
    num_syn_students = len(student_elements)
    print("Num Students: " + str(num_syn_students))
    syn_students = []
    
    time.sleep(.5)
    element = driver.find_element(By.XPATH,"//*[@id='div_scroll_Top']/span")
    if element.is_enabled() and element.is_displayed():
        driver.execute_script("return arguments[0].scrollIntoView(true);", element)
        element.send_keys(Keys.HOME)
        time.sleep(.5)
    #driver.execute_script("window.scrollTo(0,40)")      
    driver.find_element(By.TAG_NAME,'body').send_keys(Keys.PAGE_UP) 
    time.sleep(1)
    print("Scroll up to show the first student then hit ENTER if script is stalled longer than 30 seconds (you should see underlines on the student names)")
    
    #https://stackoverflow.com/questions/70116814/how-to-mouse-hover-multiple-elements-one-by-one-and-then-click-on-an-element-wit
    
    
    # https://stackoverflow.com/questions/62376755/how-to-mouse-hover-multiple-elements-using-selenium-and-extract-the-texts-throug
    # elements = WebDriverWait(driver, 5).until(EC.visibility_of_all_elements_located((By.XPATH, "//img[@class='market_listing_item_img economy_item_hoverable']")))
    # for element in elements:
    # ActionChains(driver).move_to_element(element).perform()
    # print([my_elem.text for my_elem in WebDriverWait(driver, 5).until(EC.visibility_of_all_elements_located((By.XPATH, "//div[@class='item_desc_description']//div[@class='item_desc_descriptors' and @id='hover_item_descriptors']//div[@class='descriptor']")))])
    print("Scanning Student ID's...")
    for found_student in student_elements: 
        #element.send_keys(Keys.DOWN)
        #found_student = driver.find_element(By.XPATH,"//*[@id='ctl00_cphbody_gv_StudentsScores_ctl" + format(2 + n,'02d')  + "_lbl_studentLastFirst']")       
        #need to hover so element loads student id
        #driver.execute_script("return arguments[0].scrollIntoView(true);", found_student)
        found_student_id = None

        while found_student_id == None:
            hover = ActionChains(driver).move_to_element(found_student)
            hover.perform()
            time.sleep(.05)
            found_student_id = found_student.get_attribute('data-original-title') 
            #print("Scroll up")
        
        if found_student_id != None:
            #print("Student ID: " + str(found_student_id))
            canvas_student_row = -1
            #Check to see if student has scores in csv file
            for m in range(first_csv_datarow - 1,num_csv_rows - 1):   
                #print("Checking Student: " + data[m][0])
                if data[m][2].isdigit():
                    if (int(data[m][2]) == int(found_student_id) ):
                        #Check if the section matches the student record (you may have students in multiple sections)
                        if section in data[m][4]:
                            canvas_student_row = m
                        else:
                            print("Found student in gradebook scores export but the section didn't match. The course title is probably wrong in Sections.xls. Did you update everything to 2 when the semester changed?" )
            new = []
            new.append(canvas_student_row)
            syn_students.append(new)
        else:
            print("ERROR: Student missed, stop script - student not in view")
            
    driver.find_element(By.TAG_NAME,'body').send_keys(Keys.PAGE_UP) 
    num_syn_assignments = len(driver.find_elements(By.XPATH,"//*[contains(@id,'Assignment_Header_')]"))
    print("Num Synergy Assignments: " + str(num_syn_assignments))
    syn_assignment_names = []
    
    time.sleep(1)
    element = driver.find_element(By.XPATH,"//*[@id='div_scroll_Top']/span")
    if element.is_enabled() and element.is_displayed():
        element.send_keys(Keys.HOME)
    time.sleep(1)
    
    #driver.execute_script("return arguments[0].scrollIntoView(true);", element)
    element = driver.find_element(By.TAG_NAME,'body')
    element.send_keys(Keys.PAGE_UP) 
    time.sleep(.5)
    element.send_keys(Keys.HOME) 
    time.sleep(.5)
    driver.execute_script("window.scrollTo(0,0)")   
    time.sleep(.5)
    
    for n in range(0,num_syn_assignments): 
        element = driver.find_element(By.XPATH,"//*[@id='div_scroll_Top']/span")
        if element.is_enabled() and element.is_displayed():
            driver.execute_script("return arguments[0].scrollIntoView(true);", element)
            element.send_keys(Keys.RIGHT)
            time.sleep(.25)

        #driver.find_element(By.XPATH,"//*[@id='div_scroll_Top']/span").send_keys(Keys.RIGHT)
        found_assignment_element = driver.find_element(By.XPATH,"//*[@id='Assignment_Header_" + str(n) + "']")
        #hover = ActionChains(driver).move_to_element(found_assignment_element)
        #hover.perform()

        time.sleep(.1)
        found_assignment = found_assignment_element.get_attribute("innerHTML")
        ass_name = re.sub(r'\s\([0-9]{6,}\)$', '', found_assignment ).strip()
        
        if ass_name == "":
            print("Error: An assignment was not found in Synergy so all assignments will not sync. Don't use any other applications right now to avoid this issue")
        else:
            print("Found: " + ass_name)
            syn_assignment_names.append(ass_name.strip())
    

    time.sleep(1)
    element = driver.find_element(By.XPATH,"//*[@id='div_scroll_Top']/span")
    if element.is_enabled() and element.is_displayed():
        driver.execute_script("return arguments[0].scrollIntoView(true);", element)
        element.send_keys(Keys.HOME)

    driver.find_element(By.TAG_NAME,'body').send_keys(Keys.PAGE_UP) 
    driver.find_element(By.TAG_NAME,'body').send_keys(Keys.HOME) 
    driver.execute_script("window.scrollTo(0,0)")   
    time.sleep(1)
    #Read assienment names from csv col headers and save to array
    col = 0
    num_canvas_ass = 0
    for col in range(0,num_csv_cols):
 
        #print(data[1][col])
        #print(data[0][col])
        replaced_string = data[1][col].replace(".", "", 1)
        if replaced_string.isdigit():
            assignment_name = data[0][col].strip()
            #Remove assignment number (3243243) from header
            ass_name = re.sub(r'\s\([0-9]{6,}\)$', '', assignment_name )
            new = []
            new.append(col)
            new.append(ass_name.strip())
            canvas_headings.append(new)
            #print("Adding: " + ass_name)
            num_canvas_ass = num_canvas_ass + 1
    next
    
    print("Num Canvas Assignments: " + str(num_canvas_ass))
    
    for n in range(3,0,-1):
                print("Waiting... ")
                time.sleep(1)
        
    for col in range(1,num_syn_assignments + 1):
        #driver.execute_script("window.scrollTo(0,40)")        
        driver.find_element(By.TAG_NAME,'body').send_keys(Keys.PAGE_UP) 
        time.sleep(.5)    
        print("Syncing: " + syn_assignment_names[col - 1] + " ")
        #tried moving the slider but couldn't scroll to the first student so went back to search method
        #if not dryrun and col > (1 + assignmentstoskip): #move the grades table one col right
            #element = driver.find_element(By.XPATH,"//*[@id='div_scroll_Top']/span")
            #driver.execute_script("return arguments[0].scrollIntoView(true);", element)
            #element.send_keys(Keys.RIGHT)
            #driver.find_element(By.XPATH,"//*[@id='div_scroll_Top']/span").send_keys(Keys.RIGHT)
        
        match_ratio = 0
        matched_ass = ""
        for n in range(0,num_canvas_ass): 
            #print(canvas_headings[n][1], syn_assignment_names[col - 1])
            check_ratio = fuzz.ratio(canvas_headings[n][1],syn_assignment_names[col - 1])           
            if check_ratio > match_ratio:
                match_ratio = check_ratio
                canvas_ass_col = canvas_headings[n][0]
                matched_ass = canvas_headings[n][1]
                
        #print("Canvas Col: " +  str(canvas_ass_col))
        #print("Matched : '" +  matched_ass + "' Ratio: " + str(match_ratio))

        if match_ratio > 90: #proceed only if assignment names match
            syn_assigment_name = syn_assignment_names[col - 1]

            driver.find_element(By.ID, "txt_GradeBookSearch").click()
            driver.find_element(By.ID, "txt_GradeBookSearch").send_keys(Keys.CONTROL + "a")
            driver.find_element(By.ID, "txt_GradeBookSearch").send_keys(syn_assigment_name)
            driver.find_element(By.ID, "txt_GradeBookSearch").send_keys(Keys.ENTER)
            #time.sleep(.5)

            if col > assignmentstoskip - 1:
                #loop students
                for row in range(0, num_syn_students):
                    #print("Student " + str(row) + " " + data[row - 1][0]) 
                    canvas_row = syn_students[row][0]
                    if canvas_row > 0:
                        #The actual rows on the webpage are shifted by 1 because the header is first
                        childnum = row  + 2 
                           
                        #if dryrun:
                        #print("Student " + str(row) + " " + data[row - 1][0]) #Student Name
                        
                        if (childnum % 2) == 0: #Even children .cAGR:nth-child(#)
                            element = driver.find_element(By.CSS_SELECTOR, ".cAGR:nth-child(" + str(childnum)  + ") > .cAGR:nth-child(" + str(col) + ") .SAI")
                        else:                   #Odd children cAGRA:nth-child(#)
                            element = driver.find_element(By.CSS_SELECTOR, ".cAGR:nth-child(" + str(childnum)  + ") > .cAGRA:nth-child(" + str(col) + ") .SAI")
                        #//*[@id="ctl00_cphbody_GV_Assignments"]/tbody/tr[4]/td[10]/div/img
                        #//*[@id="ctl00_cphbody_GV_Assignments"]/tbody/tr[4]/td[10]/div/div[3]/input
                        element.click()
                        curElement = driver.find_element(By.CSS_SELECTOR, ".GB_AssignmentCellEditMode .GB_AssignmentGridTextbox")
                          
                        newscore = data[canvas_row][canvas_ass_col]
                        #print("New Score: " + newscore)
                        replaced_string = newscore.replace(".", "", 1)
                        if replaced_string.isdigit():
                            tail_dot_rgx = re.compile(r'(?:(\.)|(\.\d*?[1-9]\d*?))0+(?=\b|[^0-9])')
                            newscore = tail_dot_rgx.sub(r'\2',newscore) #strip any trailing zeros or decimal points
                        
                        oldscore = curElement.get_attribute("value")
                        if  (oldscore != newscore):
                            num_updates = num_updates + 1
                            if not newscore: #check for blank score and remove any existign scores
                                newscore = Keys.BACKSPACE                       
                            if (newscore != "N/A"):
                                curElement.send_keys(newscore)
                        
                next
        else:
            print("Skipping: '" + syn_assignment_names[col - 1] + "'" + " (Not Found in Canvas Export)")
    next

    savescores(driver, num_updates ) 
    element = driver.find_element(By.XPATH,"//*[@id='div_scroll_Top']/span")
    if element.is_enabled() and element.is_displayed():
        driver.execute_script("return arguments[0].scrollIntoView(true);", element)
        element.send_keys(Keys.HOME)
    for n in range(3,0,-1):
        print("Waiting... ")
        time.sleep(1)
    
    
def sortgradebook(driver):
    print("\nMake Sure Student Sort is A-Z" )
    timeout = 5

    #driver.find_element(By.LINK_TEXT, "Filters & Options").click()
    #This fixes a stale element error when clicking the filters and options button
    click_stale_element_xpath(driver, "//*[@id='ctl00_lowerFixedBarContainer_LeftButtonContainer']")
    
    element = driver.find_element(By.ID, "ctl00_lowerFixedBarContainer_dpd_RowSize")
    click_ni_element(element)
    
    dropdown = driver.find_element(By.ID, "ctl00_lowerFixedBarContainer_dpd_RowSize")
    #time.sleep(.1)
    dropdown.find_element(By.XPATH, "//option[. = 'Small']").click()

    driver.find_element(By.ID, "ctl00_lowerFixedBarContainer_dpd_StudentSort").click()
    dropdown = driver.find_element(By.ID, "ctl00_lowerFixedBarContainer_dpd_StudentSort")
    #time.sleep(.1)
    dropdown.find_element(By.XPATH, "//option[. = 'Last Name']").click()

    driver.find_element(By.ID, "ctl00_lowerFixedBarContainer_dpd_AssignmentDateFilter").click()
    dropdown = driver.find_element(By.ID, "ctl00_lowerFixedBarContainer_dpd_AssignmentDateFilter")
    #time.sleep(.1)
    dropdown.find_element(By.XPATH, "//option[. = 'All Dates']").click()

    driver.find_element(By.ID,  "ctl00_lowerFixedBarContainer_dpd_AssignmentOrder").click()
    dropdown = driver.find_element(By.ID, "ctl00_lowerFixedBarContainer_dpd_AssignmentOrder")
    #time.sleep(.1)
    dropdown.find_element(By.XPATH, "//option[. = 'Oldest to Newest']").click()

    driver.find_element(By.ID,  "ctl00_lowerFixedBarContainer_dpd_AssignmentFilter").click()
    dropdown = driver.find_element(By.ID, "ctl00_lowerFixedBarContainer_dpd_AssignmentFilter")
    #time.sleep(.1)
    dropdown.find_element(By.XPATH, "//option[. = 'Show All']").click()
    time.sleep(1)
    driver.find_element(By.LINK_TEXT, "DONE").click()
    #<input type="image" name="ctl00$cphbody$gv_StudentsScores$ctl01$btn_StudentSort" id="ctl00_cphbody_gv_StudentsScores_ctl01_btn_StudentSort" title="" class="img_Sort" data-toggle="tooltip" lockondirtypage="true" src="resources/GBResources/GB_SortUp_Over.png?2021-02-12T17%3a32%3a31.0000000" style="border-width:0px;" data-original-title="Sort Direction">

    sort_icon_el = driver.find_element(By.XPATH,"//*[@id='ctl00_cphbody_gv_StudentsScores_ctl01_btn_StudentSort']")
    sort_icon_src = sort_icon_el.get_attribute("src")
    print(sort_icon_src)
       
    if not ("GB_SortUp_Over.png" in sort_icon_src): 
        click_stale_element(sort_icon_el)
        
    #for n in range(1,0,-1):
        #print("Waiting... " + str(n))
        #speedup time.sleep(1)
    
    
    
    
def createassignmentsincourse(driver, course_code, ass_group_data, othersectionid1 = "",othersectionid2 = "",othersectionid3 = "",othersectionid4 = "",othersectionid5 = ""):
    
    SIS_COURSE_CODE = 0
    COURSE_CODE = 1
    WORKFLOW_STATE = 2
    COURSE_NAME = 3
    ASS_ID = 4
    ASSIGNMENT_NAME = 5
    PUBLISHED = 6
    POINTS_POSSIBLE = 7
    DESCRIPTION = 8
    DUE_AT = 9
    LOCK_AT = 10
    UNLOCK_AT = 11
    SUBMISSION_TYPES = 12
    GRADING_TYPE = 13
    GROUP_CATEGORY_ID = 14
    HAS_OVERRIDES = 15
    COURSE_ID = 16
    CREATED_AT = 17
    UPDATED_AT = 18
    HTML_URL = 19
    SUBMISSIONS_DOWNLOAD_URL = 20
    ASSIGNMENT_GROUP_ID = 21
    DUE_DATE_REQUIRED = 22
    MAX_NAME_LENGTH = 23
    TURNITIN_ENABLED = 24
    VERICITE_ENABLED = 25
    
    #Section Info Excel file Assignment Groups Tab (0 based) 
    ASS_GROUP_SYNC = 1
    ASS_GROUP_NAME = 3
    ASS_SHOW_ONLY_WHEN_SCORED = 4

    MIN_MATCH_RATIO = 90
    
    wb = load_workbook(filename = 'CanvasApiAssignmentsReport1v6.xlsm')
    firstdatarow = 5 
    datacols = 25

    sheet_ranges = wb['Main report']
    n = 1
    print("Scanning Synergy Gradebook")
    time.sleep(1)
    #Get number of assinments
    #get_xpath_count("xpath=//div*[matches(text(), \"day \\d night\")]");
    num_syn_assignments = len(driver.find_elements(By.XPATH,"//*[contains(@id,'Assignment_Header_')]"))
    #print("Num Synergy Assignments: " + str(num_syn_assignments))   
    
    syn_assignment_names = []
    for n in range(0,num_syn_assignments): 
        element = driver.find_element(By.XPATH,"//*[@id='div_scroll_Top']/span")
        if element.is_enabled() and element.is_displayed():
            driver.execute_script("return arguments[0].scrollIntoView(true);", element)
            element.send_keys(Keys.RIGHT)
            time.sleep(.25)
   
        #found_assignment = driver.find_element(By.XPATH,"//*[@id='Assignment_Header_" + str(n) + "']").text
        #driver.find_element(By.XPATH,"//*[@id='div_scroll_Top']/span").send_keys(Keys.RIGHT)
        
        
        found_assignment_element = driver.find_element(By.XPATH,"//*[@id='Assignment_Header_" + str(n) + "']")
        #hover = ActionChains(driver).move_to_element(found_assignment_element)
        #hover.perform()
        time.sleep(.1)
        if found_assignment_element:
            found_assignment = found_assignment_element.get_attribute("innerHTML").strip()
            
            if found_assignment == "":
                print("Error: An assignment was not found in Synergy so all assignments will not sync. Don't use any other applications right now to avoid this issue")
            else:
                print("Found " + found_assignment + " in Synergy")
                syn_assignment_names.append(found_assignment)

    element = driver.find_element(By.XPATH,"//*[@id='div_scroll_Top']/span")    
    if element.is_enabled() and element.is_displayed():
        element.send_keys(Keys.HOME)
        time.sleep(.5)

    #//*[@id='Assignment_Header_
    #//*[contains(@id,'Assignment_Header_')]
    #IMPORTANT DON'T FORGET -- Assignments in the "Imported Assignments" group are skipped unless that type is in the Sections spreadsheet
    for line in sheet_ranges.iter_rows(min_row=firstdatarow, max_col=datacols):

        if line[COURSE_CODE].value == course_code:
            print("Processing: " + line[ASSIGNMENT_NAME].value)
            match_ratio = 0
            #print("Searching: '" + line[5].value + "'")
            for n in range(0,num_syn_assignments): 
                ass_name = re.sub(r'\s\([0-9]{6,}\)$', '', syn_assignment_names[n] ) #Remove ass id from syn name if exists
                check_ratio = fuzz.ratio(ass_name,line[ASSIGNMENT_NAME].value.strip())
                #print(check_ratio)
                if check_ratio > match_ratio:
                    match_ratio = check_ratio
                    assignment_name = syn_assignment_names[n]
            
            #print("Hightest Match: " + str(match_ratio) + " Name: '" + assignment_name + "'")

            #element = driver.find_element(By.ID, "ctl00_lowerFixedBarContainer_btn_NewAssignment")
            ind = np.where(np.array(ass_group_data) == str(line[ASSIGNMENT_GROUP_ID].value))
            # print(str(line[ASSIGNMENT_GROUP_ID]))
            # print(ind)
            # print(ass_group_data[ind[0]][ind[1]])
            # print(ass_group_data[ind[0]][ind[1]][0][3] )
            
            
            if ass_group_data[ind[0]][ind[1]].size == 0:
                print("Assignment group " + str(line[ASSIGNMENT_GROUP_ID].value) + " missing! Add it to your Sectioninfo.xlsx spreadsheet if you want to sync this assignment group. If this seems to be an error and the types are present be sure all cells are formatted as text")
                sync_ass = "No"
                ass_type_name = "None"
            else:
                ass_type_name =         ass_group_data[ind[0]][ind[1]][0][ASS_GROUP_NAME]
                sync_ass =              ass_group_data[ind[0]][ind[1]][0][ASS_GROUP_SYNC]
                showonlywhenscored =    ass_group_data[ind[0]][ind[1]][0][ASS_SHOW_ONLY_WHEN_SCORED]

            if (line[DUE_AT].value != None) and (match_ratio < MIN_MATCH_RATIO) and (line[POINTS_POSSIBLE].value > 0) and (line[PUBLISHED].value > 0) and (sync_ass == "Yes"): #Skip assignmts without date and there are matches and have score and are published and don't have synced types

                #try this next: 
                #https://sqa.stackexchange.com/questions/40678/using-python-selenium-not-able-to-perform-click-operation
                print("Adding...")
                driver.execute_script("javascript:__doPostBack('ctl00$lowerFixedBarContainer$btn_NewAssignment','')")
                #speedup time.sleep(2)
        
                #speedup time.sleep(1)
                
                driver.find_element(By.ID, "ctl00_cphbody_txtMeasure").send_keys(line[ASSIGNMENT_NAME].value + " (" + line[ASS_ID].value + ")") #Name with canvas course in parenthesis
                driver.find_element(By.ID, "ctl00_cphbody_txtDescription").send_keys(line[HTML_URL].value)

                driver.find_element(By.ID, "ctl00_cphbody_txtMaxValue").click()
                driver.find_element(By.ID, "ctl00_cphbody_txtMaxValue").send_keys(str(line[POINTS_POSSIBLE].value))
                driver.find_element(By.ID, "ctl00_cphbody_txtWeight").send_keys(Keys.CONTROL + "a")
                driver.find_element(By.ID, "ctl00_cphbody_txtWeight").send_keys(str(line[POINTS_POSSIBLE].value))
                #speedup time.sleep(.5)
                
                
                #https://stackoverflow.com/questions/25823608/find-matching-rows-in-2-dimensional-numpy-array
                
             
                dropdown = Select(driver.find_element(By.ID, "ctl00_cphbody_dropMeasureType"))
                dropdown.select_by_visible_text(ass_type_name)
                #one liner version of above 2 lines driver.find_element(By.XPATH,"//select[@id='ctl00_cphbody_dropMeasureType']/option[text()='Assignments']").click()
                #old - not working dropdown.find_element(By.XPATH, "//option[. = 'Assignments']").click()
                element_id = "ctl00_cphbody_calMeasureDate_I"
                time.sleep(.5)
                driver.execute_script("document.getElementById('" + element_id + "').focus();")
                time.sleep(.5)
                element = driver.find_element(By.ID, element_id)
                element.click()
                time.sleep(.5)
                element.send_keys(Keys.CONTROL + "a")
                #create_date = datetime.datetime.strptime(line[17], "%d %b %Y %H:%M").strftime('%m/%d/%y')
                create_date = line[CREATED_AT].value.date().strftime('%m/%d/%Y')
                #action = ActionChains(driver)
                #action.move_to_element(element).perform();
                #driver.execute_script("document.getElementById('ctl00_cphbody_calMeasureDate_I').setAttribute('value', '" + create_date + "');")
                #driver.execute_script("document.getElementById('ctl00_cphbody_calMeasureDate_I').value = '" + create_date + "');")
                time.sleep(.5)
                element.send_keys(create_date) #Date Created
                time.sleep(.5)
                
                element_id = "ctl00_cphbody_calDueDate_I"
                driver.execute_script("document.getElementById('" + element_id + "').focus();")
                time.sleep(.5)
                element = driver.find_element(By.ID, element_id)
                element.click()
                time.sleep(.5)
                element.send_keys(Keys.CONTROL + "a")
                time.sleep(.5)
                #print("Type: " + str(type(line[9].value)))
                if line[DUE_AT].is_date:
                    due_date = line[DUE_AT].value.date().strftime('%m/%d/%Y')
                else:
                    #if multiple due dates, parse array
                    duedatesstr = re.search(r'{([^}]*)}', line[DUE_AT].value)
              
                    for datestr in duedatesstr.group(1).split(","):
                        due_date = datetime.datetime.strptime(datestr[1:-1], "%d %B %Y %H:%M").strftime('%m/%d/%Y')
                        
                element.send_keys(due_date)
                #action = ActionChains(driver)
                #action.move_to_element(element).perform();
                #driver.execute_script("document.getElementById('ctl00_cphbody_calDueDate_I').setAttribute('value','" + due_date + "');")
               
                time.sleep(.5)
                if showonlywhenscored == "Yes":
                    driver.find_element(By.XPATH,"//*[@id='ctl00_cphbody_div_ShowWhenScoreEntered']/div[2]/div").click()
   
                time.sleep(.5)
                
                wait_for_element_to_load(driver,"ctl00_cphbody_ul_TabsActiveTab")
                element = driver.find_element(By.ID, "ctl00_cphbody_ul_TabsActiveTab")
                driver.execute_script("return arguments[0].scrollIntoView(true);", element)
                time.sleep(.5)
                element = driver.find_element(By.LINK_TEXT, "Sections")
                # driver.execute_script("return arguments[0].scrollIntoView(true);", element)
                # action = ActionChains(driver)
                #driver.find_element(By.TAG_NAME,'body').send_keys(Keys.PAGE_UP) 
                driver.find_element(By.TAG_NAME,'body').send_keys(Keys.PAGE_DOWN) 
                # action.move_to_element(element).perform()
            
                time.sleep(.5)
                element.click()
                
                if othersectionid1 != "":
                    #driver.find_element(By.LINK_TEXT, "Sections").click()
                    element = driver.find_element(By.XPATH,"//*[ contains (text(),'SEC:" + othersectionid1 + "' ) ]")
                    driver.execute_script("return arguments[0].scrollIntoView(true);", element)
                    element.click()
                    
                if othersectionid2 != "":
                    #driver.find_element(By.LINK_TEXT, "Sections").click()
                    element = driver.find_element(By.XPATH,"//*[ contains (text(),'SEC:" + othersectionid2 + "' ) ]")
                    driver.execute_script("return arguments[0].scrollIntoView(true);", element)
                    element.click()
                
                if othersectionid3 != "":
                    #driver.find_element(By.LINK_TEXT, "Sections").click()
                    element = driver.find_element(By.XPATH,"//*[ contains (text(),'SEC:" + othersectionid3 + "' ) ]")
                    driver.execute_script("return arguments[0].scrollIntoView(true);", element)
                    element.click()          

                if othersectionid4 != "":
                    #driver.find_element(By.LINK_TEXT, "Sections").click()
                    element = driver.find_element(By.XPATH,"//*[ contains (text(),'SEC:" + othersectionid4 + "' ) ]")
                    driver.execute_script("return arguments[0].scrollIntoView(true);", element)
                    element.click()     

                if othersectionid5 != "":
                    #driver.find_element(By.LINK_TEXT, "Sections").click()
                    element = driver.find_element(By.XPATH,"//*[ contains (text(),'SEC:" + othersectionid5 + "' ) ]")
                    driver.execute_script("return arguments[0].scrollIntoView(true);", element)
                    element.click()     


                #Return to top after scorlling to other sections

                driver.find_element(By.TAG_NAME,'body').send_keys(Keys.CONTROL + Keys.HOME)  
                
                driver.find_element(By.TAG_NAME,'body').send_keys(Keys.PAGE_UP) 
                driver.find_element(By.TAG_NAME,'body').send_keys(Keys.PAGE_DOWN) 
                #Select Grading Periods
                wait_for_element_to_load(driver,"ctl00_cphbody_ul_TabsActiveTab")
                element = driver.find_element(By.ID, "ctl00_cphbody_ul_TabsActiveTab")
                #driver.execute_script("return arguments[0].scrollIntoView(true);", element)
                time.sleep(.5)
                element = driver.find_element(By.LINK_TEXT, "Grading Periods")
                element.click()
                
                numperiods = len(driver.find_elements(By.XPATH,"//*[contains(@id,'ctl00_cphbody_chkPeriods_')]"))
                
                for n in range(0,numperiods): 
                    element = driver.find_element(By.XPATH,"//*[@id='ctl00_cphbody_chkPeriods_" + str(n) + "']")
                    driver.execute_script("return arguments[0].scrollIntoView(true);", element)
                    if not element.get_attribute("checked"):
                        element.click()
                        
                driver.find_element(By.TAG_NAME,'body').send_keys(Keys.CONTROL + Keys.HOME) 
                time.sleep(.5)
                #for n in range(5,0,-1):
                #    print("Saving... " + str(n))
                #    time.sleep(1)
                
                with wait_for_page_load(driver,timeout=10):
                    print("Saving... ")
                    driver.execute_script("javascript:WebForm_DoPostBackWithOptions(new WebForm_PostBackOptions('ctl00$lowerFixedBarContainer$btn_Save', '', true, 'Save', '', false, true))")
                    handleAlert(driver)
            else:
                if match_ratio > MIN_MATCH_RATIO:
                    print("Skipping... Assignment already exists, match ratio: " + str(match_ratio))
                else:
                    print("Skipping...")
                    print("Sync: " + ass_type_name + " " + sync_ass)
                    print("Due: " + str(line[DUE_AT].value == None) )
                    print("Match: " + str(match_ratio) + " Min: " + str(MIN_MATCH_RATIO))
                    print("Points: " + str(line[POINTS_POSSIBLE].value))
                    print("Published: " + str(line[PUBLISHED].value))
                
          


def changesection(driver,syn_sectionid,syn_course_title):
    print("\nChanging Synergy section...\n")
    global current_syn_section_id 
    #print(syn_sectionid)
    if (syn_sectionid != current_syn_section_id):
        #for n in range(3,0,-1):
            #print("Waiting for page load... " + str(n))
            #speedup time.sleep(1)
        wait_for_element_to_load(driver, "lbl_FocusButton", 4)
        
        #(S1) Fenger, N  Engineering 3(3) SEC:1695B1-F3 / Q1 Progress
        #more here: https://stackoverflow.com/questions/14590341/use-python-selenium-to-get-span-text
        selected_section = driver.find_element(By.XPATH,"//*[@id='lbl_FocusButton']").get_attribute("innerHTML")
        
        if selected_section.find(syn_sectionid) < 0:
           
            #element_to_hover_over = driver.find_element(By.XPATH,"//*[@id="lbl_FocusButton")
            element_to_hover_over = driver.find_element(By.ID, "lbl_FocusButton")     
            hover = ActionChains(driver).move_to_element(element_to_hover_over)
            hover.perform()
            #for n in range(3,0,-1):
            #    print("Waiting for hover menu... " + str(n))
                #speedup time.sleep(1)
            wait_for_element_to_load(driver, "focus_selections", 30)
            #element = driver.find_element(By.ID, "lbl_FocusButton")
            #click_ni_element(element)
            print("Selecting " + syn_course_title + " " + syn_sectionid)
            #Note - we are getting lucky here with T/A sections 
            #as they have the same section id as the main section so we may have to 
            #add some code if they aren't first in the list
            #//*[@id="HomeRoomClasses"]/table/tbody/tr[5]/td[2]/div[2]
            #HomeRoomClasses > table > tbody > tr:nth-child(5) > td:nth-child(2) > div:nth-child(2)
            #driver.find_element(By.XPATH,"//*[ contains (text(),'" + syn_sectionid + "' ) ]").click()

            driver.find_element(By.XPATH,"//table[@class='table-focus']/tbody/tr[td[contains(text(), '" + syn_sectionid + "')] and td[div[text()='" + syn_course_title + "' ] ]]").click()



            for n in range(5,0,-1):
                print("Waiting for section to change... ")
                time.sleep(1)
                selected_section = driver.find_element(By.XPATH,"//*[@id='lbl_FocusButton']").get_attribute("innerHTML")
                if selected_section.find(syn_sectionid) != -1:
                    break
            print("Changed section to: " + selected_section)
            
            #//table[@class='table-focus']/tbody/tr[td[contains(text(), '1695B2-F3') ]  ]
            #//table[@class='table-focus']/tbody/tr[td/div[contains(text(), 'Engineering 4') ]  ]
            #https://www.google.com/search?q=selenium+find+table+row+by+text+multiple+conditions+pythonn
#https://www.youtube.com/watch?v=OTStvDR_jF4
#https://stackoverflow.com/questions/19721111/how-to-search-node-by-exact-text-match-using-xpath-in-webdriver
#https://stackoverflow.com/questions/4037255/selenium-xpath-to-match-a-table-row-containing-multiple-elements
#https://stackoverflow.com/questions/65841484/python-selenium-find-element-by-xpath-multiple-conditions/65841771
#https://stackoverflow.com/questions/56817313/python-selenium-how-to-find-an-element-not-containing-a-string
 #           for n in range(5,0,-1):
  #              print("Verify T/A Section Wasn't Selected... ")
   #             time.sleep(1)


        current_syn_section_id = syn_sectionid
    else:
        print("\nCorrect section already selected.\n")
   
def getcsvfilename(searchstring):
    csvfilename = ""
    for f in glob.glob('*' + searchstring + '*.csv'):
        csvfilename = f
    if not (searchstring in csvfilename):
        csvfilename = ""
    return csvfilename

def savescores(driver, num_updates):
    
    for n in range(SCORE_SAVE_DELAY,0,-1):
        print("Waiting... " + str(n))
        time.sleep(1)
    if num_updates > 0:
        with wait_for_page_load(driver,timeout=(10 + num_updates/20)):
            print("Saving Scores...")
            driver.find_element(By.XPATH,"//*[@id='ctl00_lowerFixedBarContainer_SaveAssignmentsButtonContainer']").click()
        print("Scores Saved!")
    else:
        print("No Changes")
    #for n in range(int(wait_time),0,-1):
    #    print("Saving... " + str(n))
    #    time.sleep(1)

def runscoreupdate(driver,synergy_section_id,synergy_course_title,canvas_section,csvfilename,assignmentstoskip):

    changesection(driver,synergy_section_id,synergy_course_title)
    #sortgradebook(driver)
    print("Running Score Update from File: " + csvfilename + " for " + synergy_section_id )
    addscores(driver, csvfilename, canvas_section,assignmentstoskip)
     
def launchchrome():
    options = webdriver.ChromeOptions()
    options.add_argument("start-maximized")
    options.add_argument("disable-infobars")
    options.add_argument("disable-extensions")
    #options.add_argument("headless")
    #driver = webdriver.Chrome(options=options, executable_path=r'C:\Utility\BrowserDrivers\chromedriver.exe')
    driver = webdriver.Chrome(options=options)
    return driver

def launchsynergy():
    password = ""
    syn_username = ""
    home = str(Path.home())
    print(home)
    
    if socket.gethostname().find('.')>=0:
        machine_name=socket.gethostname()
    else:
        machine_name=socket.gethostbyaddr(socket.gethostname())[0]
    print("machine:" + machine_name)
    
    domain = os.environ['userdomain']
    win_username = getpass.getuser()
    
    print("domain:" + domain)
    

    if args.username is None:
        if domain == "PPS":
            syn_username = win_username
        elif domain == "IT00-208821":
            syn_username = win_username
    else:
        syn_username = args.username
        
        
    if os.path.exists(home + '\synergypass.encrypted') and os.path.exists('decrypt.key.' + machine_name):

        with open('decrypt.key.' + machine_name, 'rb') as keyfile:
            key = keyfile.read()
        with open(home + '\synergypass.encrypted', 'rb') as passfile:
            encoded_password = passfile.read()
        
        cipher_suite = Fernet(key)
        
        config = configparser.ConfigParser()
        password = cipher_suite.decrypt(encoded_password).decode()  
    else:

        password = pwinput.pwinput(prompt='Enter your synergy passsword: ')
        key = Fernet.generate_key() 
       
        with open('decrypt.key.' + machine_name, 'wb') as keyfile:
            keyfile.write(key)
           
        cipher_suite = Fernet(key)
        encoded_password = cipher_suite.encrypt(password.encode())

        with open(home  + '/synergypass.encrypted', 'wb') as passfile:
            passfile.write(encoded_password)
 
    driver = launchchrome()
    #driver.set_window_size(1500, 1100)

    #driver.get('chrome://settings/')
    #driver.execute_script('chrome.settingsPrivate.setDefaultZoom(.5);')
    #driver.execute_script("document.body.style.zoom='50
    driver.get("https://sis-portland.cascadetech.org/Portland/Login.aspx")
    
    if len(password) > 0:   
        driver.find_element(By.ID, "login_name").click()
        driver.find_element(By.ID, "login_name").send_keys(syn_username)
        
        driver.find_element(By.NAME, "password").send_keys(password)
        driver.find_element(By.NAME, "btnLogin").click()

    else:
        for n in range(30,0,-1):
            print("Login..." + str(n))
            time.sleep(1)
    
    driver.switch_to.frame(0)
    
    return driver

def sync():
    #Section Info Spreadssheet Columns
    SEC_TITLE =     0
    SEC_SYNC =      1
    SEC_SKIP =      2
    #SEC_SHOW =      3  #not used 
    SEC_COURSE_CODE = 4
    SEC_PREFIX =    5
    SEC_PARENT =    6
    SEC_PA_NUM =    7
    SEC_PA_NAME =   8
    SEC_PB_NUM =    9
    SEC_PB_NAME =   10
    SEC_PC_NUM =    11
    SEC_PC_NAME =   12
    SEC_PD_NUM =    13
    SEC_PD_NAME =   14
    SEC_PE_NUM =    15
    SEC_PE_NAME =   16
    SEC_PF_NUM =    17
    SEC_PF_NAME =   18
    
    driver = launchsynergy()
    #driver = None

    driver.get("https://sis-portland.cascadetech.org/Portland/gb_GradeBookMain.aspx")
    driver.implicitly_wait(10)
    
    for n in range(1,0,-1):
        #print("Select Correct Term... " + str(n))
        time.sleep(1)
    print("Loading 'Section Info' spreadsheet ")
    wb = load_workbook(filename = 'Sections.xlsx')
    maxdatacols = 25
    #print(wb.sheetnames)
    sheet_ranges = wb['Section Info']       
    print("Reading spreadsheet...")
    sline = 0
    for secline in sheet_ranges.iter_rows(min_row=2, max_col=maxdatacols, values_only=True):
        print("Reading assignment group data...")
        ass_group_sheet = wb['Assignment Groups']        
        ass_group_data = np.array([[cell.value for cell in row] for row in ass_group_sheet.iter_rows(min_row=2, max_col=5 )])
        # ind = np.where(np.array(ass_group_data) == '3412312')
        # print(ind)
        # print(ass_group_data[ind[0]][ind[1]])
        
        #print(ass_group_data)        
        #print(secline)
        #print(secline[SEC_SYNC])
        
        
        if secline[SEC_SYNC] != None:
            course_code = secline[SEC_COURSE_CODE] 
            print("Sync " + course_code + " = " + secline[SEC_SYNC])
            if secline[SEC_SYNC] == "Yes":
                othersectionid1 = ""
                othersectionid2 = ""
                othersectionid3 = ""
                othersectionid4 = ""
                othersectionid5 = ""
               
                
                #re.sub(r'(^P)', 'F', secline[2])
                if secline[SEC_PB_NUM] != None:
                    othersectionid1 = course_code + "-" + secline[SEC_PREFIX] + str(secline[SEC_PB_NUM])
                if secline[SEC_PC_NUM] != None:
                    othersectionid2 = course_code + "-" + secline[SEC_PREFIX] + str(secline[SEC_PC_NUM])
                if secline[SEC_PD_NUM] != None:
                    othersectionid3 = course_code + "-" + secline[SEC_PREFIX] + str(secline[SEC_PD_NUM])
                if secline[SEC_PE_NUM] != None:
                    othersectionid4 = course_code + "-" + secline[SEC_PREFIX] + str(secline[SEC_PE_NUM])      
                if secline[SEC_PF_NUM] != None:
                    othersectionid5 = course_code + "-" + secline[SEC_PREFIX] + str(secline[SEC_PF_NUM])
                
                synergy_course_title = secline[SEC_TITLE]              
                regex = re.compile('[^\w ]')
                #Remove all non text from course title 
                canvas_course_title = regex.sub('', synergy_course_title)
                
                
                synergy_section_id = course_code + "-" + secline[SEC_PREFIX] + str(secline[SEC_PA_NUM])
                changesection(driver,synergy_section_id,synergy_course_title)
                print("Checking for Missing Assignments in " + canvas_course_title)
                
       
                
                createassignmentsincourse(driver, secline[SEC_COURSE_CODE] + secline[SEC_PARENT], ass_group_data, othersectionid1,othersectionid2,othersectionid3,othersectionid4,othersectionid5)
               
                
                csvfilename = getcsvfilename(secline[SEC_COURSE_CODE])

                if len(csvfilename) > 0:

                    canvas_section = canvas_course_title + secline[SEC_PA_NAME]
                    print("Updating Scores for: " + canvas_section)
                    
                    runscoreupdate(driver,synergy_section_id,synergy_course_title, canvas_section,csvfilename,secline[SEC_SKIP])
                    
                    if secline[SEC_PB_NUM] != None: #other section 1
                        synergy_section_id =  course_code + "-"  + secline[SEC_PREFIX] + str(secline[SEC_PB_NUM])
                        canvas_section = canvas_course_title + secline[SEC_PB_NAME]
                        print("Updating Scores for: " + canvas_section)
                        runscoreupdate(driver,synergy_section_id,synergy_course_title,canvas_section,csvfilename,secline[SEC_SKIP])
                    if secline[SEC_PC_NUM] != None: #other section 2
                        synergy_section_id =  course_code + "-"  + secline[SEC_PREFIX] + str(secline[SEC_PC_NUM])                       
                        canvas_section = canvas_course_title + secline[SEC_PC_NAME]
                        print("Updating Scores for: " + canvas_section)
                        runscoreupdate(driver,synergy_section_id,synergy_course_title,canvas_section,csvfilename,secline[SEC_SKIP])
                    if secline[SEC_PD_NUM] != None: #other section 3
                        synergy_section_id =  course_code + "-"  + secline[SEC_PREFIX] + str(secline[SEC_PD_NUM])
                        canvas_section = canvas_course_title + secline[SEC_PD_NAME]
                        print("Updating Scores for: " + canvas_section)
                        runscoreupdate(driver,synergy_section_id,synergy_course_title,canvas_section,csvfilename,secline[SEC_SKIP])
                    if secline[SEC_PE_NUM] != None: #other section 4
                        synergy_section_id =  course_code + "-"  + secline[SEC_PREFIX] + str(secline[SEC_PE_NUM])
                        canvas_section = canvas_course_title + secline[SEC_PE_NAME]
                        print("Updating Scores for: " + canvas_section)
                        runscoreupdate(driver,synergy_section_id,synergy_course_title,canvas_section,csvfilename,secline[SEC_SKIP])
                    if secline[SEC_PF_NUM] != None: #other section 5
                        synergy_section_id =  course_code + "-"  + secline[SEC_PREFIX] + str(secline[SEC_PF_NUM])
                        canvas_section = canvas_course_title + secline[SEC_PF_NAME]
                        print("Updating Scores for: " + canvas_section)
                        runscoreupdate(driver,synergy_section_id,synergy_course_title,canvas_section,csvfilename,secline[SEC_SKIP])    
                    print("Sync successful! Deleting " + csvfilename )
                    deletefile(csvfilename)  
                sline = sline + 1
                  
 #https://stackoverflow.com/questions/7732125/clear-text-from-textarea-with-selenium#:~:text=or-,webElement.,this%20will%20clear%20the%20value.&text=We%20are%20sending%20Ctrl%20%2B%20Backspace,also%20replace%20backspace%20with%20delete.

#New Machines setup
#Python:  https://www.python.org/downloads/ (include pip, add python to enviromnemt variable)
#run this commands at cpommand prompt:
#python -m pip install --upgrade pip
#pip install selenium openpyxl fuzzywuzzy numpy cryptography pwinput --upgrade


#Error Message: session not created: This version of ChromeDriver only supports Chrome version 86
#Download latest webdriver from here: https://chromedriver.chromium.org/downloads
#Git: https://git-scm.com/download/win (Accept all defaults except for "Configuring the terminal eumulator to use with Git Bash" Choose "Windows' default console window"
#python sync_synergy -p "your password here in the quotes"

#Main Program
#Get this from Synergy gradbook when you first log in - it's the default
global args
current_syn_section_id = ""
warnings.filterwarnings("ignore", category=UserWarning) 
#warnings.filterwarnings("ignore")

parser = ArgumentParser()
parser.add_argument("-u", "--username", dest="username", help="Username")
#parser.add_argument("-p", "--password", dest="password", 
#                     help="Password (in double quotes)")

args = parser.parse_args()

sync()



#Things to do
#find active section (read lbl_FocusButton)
#error checking when number of assingmens don't match
#add canvas id back to assignment
#find assignment name or id (just in case they are out of order)
#verify student name is correct (read names first) 



#https://stackoverflow.com/questions/44119081/how-do-you-fix-the-element-not-interactable-exception
#https://selenium-python.readthedocs.io/locating-elements.html
#https://stackoverflow.com/questions/48139676/how-to-get-the-value-of-an-element-in-python-selenium/48139708


# try:
    # element_present = expected_conditions.presence_of_element_located((By.XPATH,"//*[@id='ctl00_lowerFixedBarContainer_LeftButtonContainer']/li[1]/div/a" ))
    # WebDriverWait(driver, timeout).until(element_present)
# except TimeoutException:
    # print("Timed out waiting for page to load")
    
    
# my_element_id = "//*[@id='ctl00_lowerFixedBarContainer_LeftButtonContainer']/li[1]/div/a"
# ignored_exceptions=(NoSuchElementException,StaleElementReferenceException)
# your_element = WebDriverWait(driver, timeout,ignored_exceptions=ignored_exceptions)\
# .until(expected_conditions.presence_of_element_located((By.XPATH, my_element_id)))
# your_element.click()

# def find(driver):
    # my_element_id = "//*[@id='ctl00_lowerFixedBarContainer_LeftButtonContainer']/li[1]/div/a"
    # element = driver.find_elements(By.XPATH,my_element_id)
    # if element.is_enabled() and element.is_displayed():
        # return element
    # else:
        # return False
# element = WebDriverWait(driver, timeout).until(find)

# element[0].click()

        # if last_assreport_coursecode != line[0]:
            # last_assreport_coursecode = line[0]
            
            # sis_parse = re.search(r'(?<=\-)([a-zA-Z0-9]*)-(.{1})(.{1})', last_assreport_coursecode)
            # course = sis_parse.group(1)
            
            # with open(sectioninfocsvfilename, "r") as sectionf:
                
                # sectionreader = csv.reader(sectionf, delimiter=",")
                # sline = 0
                
                # for secline in sectionreader:
                    # if secline[SEC_COURSE_CODE] == course:
                        # #Change section
                     

                    # sline = sline + 1



                #wait_for_element_to_load(driver,"ctl00_cphbody_calMeasureDate_I")
                
                #wait_for_element_to_load(driver,"ctl00_cphbody_calMeasureDate_I")

                #my_element_id = 'ctl00_cphbody_calMeasureDate_I'
                #my_element_id = 'ctl00_cphbody_calMeasureDate_B-1Img'
                #wait_for_element_to_load(driver,my_element_id)
                #ignored_exceptions=(NoSuchElementException,StaleElementReferenceException,ElementNotInteractableException)
                #element = WebDriverWait(driver, 10,ignored_exceptions=ignored_exceptions)\
                #        .until(expected_conditions.presence_of_element_located((By.ID, my_element_id)))

                #wait = WebDriverWait(driver, 30);
                #element = driver.find_element(By.ID, my_element_id)
                #wait.until(EC.element_to_be_clickable((By.ID, my_element_id)))
                #element.click()
                #driver.implicitly_wait(1) 
                #ASPx.ETextChanged('ctl00_cphbody_calDueDate')
                #driver.execute_script("return ASPx.DDMC_MD('ctl00_cphbody_calMeasureDate', event)")
                #element.clear()
                #my_element_id = 'ctl00_cphbody_calMeasureDate_I'
                #wait = WebDriverWait(driver, 30);
                #element = driver.find_element(By.ID, my_element_id)
                #wait.until(EC.staleness_of(element));